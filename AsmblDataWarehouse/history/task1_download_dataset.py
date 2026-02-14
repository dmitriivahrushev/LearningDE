import time
import os
import csv
from datetime import datetime as dt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from dotenv import load_dotenv
from openpyxl import load_workbook


"""Импорт password и path для авторизации."""
ODDO_URL = os.getenv('ODDO_URL')
ODDO_LOGIN = os.getenv('ODDO_LOGIN')
ODOO_PASSWORD = os.getenv('ODOO_PASSWORD')

PATH_TO_XLSX = os.getenv('PATH_TO_XLSX')
PATH_TO_CSV = os.getenv('PATH_TO_CSV') 


def del_files():
    """Удаление датасета прошлой недели."""
    files_XLSX = os.listdir(PATH_TO_XLSX)
    for file_XLSX in files_XLSX:
        if file_XLSX.endswith('.xlsx'):
            full_file_path = os.path.join(PATH_TO_XLSX, file_XLSX)
            try:
                os.remove(full_file_path)
                print(f'Файл успешно удалён {file_XLSX}')
            except Exception as e:
                print('Произошла ошибка удаления: {e}')
    
    # Удаление raw_data.csv.
    try:
        os.remove(PATH_TO_CSV)
        print(f'Файл успешно удалён: {PATH_TO_CSV}')
    except Exception as e:
        print('Произошла ошибка удаления: {e}')
    

def load_dataset():
    """Загрузка Датасета из Odoo .XLSX"""
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(URL)
    actions = ActionChains(driver)
    time.sleep(2)

    find_login = driver.find_element(By.ID, 'login')
    find_password = driver.find_element(By.ID, 'password')
    find_login.send_keys(ODDO_LOGIN)
    find_password.send_keys(ODOO_PASSWORD)
    time.sleep(2)

    driver.find_element(By.XPATH, '//button[.="Войти"]').click()
    time.sleep(5)
    driver.find_element(By.CLASS_NAME, 'dropdown-toggle').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//a[.="Производство"]').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//span[.="Избранное"]').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//span[.="Выработка производства"]').click() 
    time.sleep(5)
    filter = driver.find_element(By.XPATH, '//span[.="Фильтры"]')
    actions.move_to_element(filter).perform()
    time.sleep(5)
    driver.find_element(By.XPATH, '//span[.="Произведено за сегодня"]').click()
    #driver.find_element(By.XPATH, '//span[.="Произведено за два дня"]').click()  
    time.sleep(5)
    driver.find_element(By.XPATH, '//button[@class="btn btn-light o_switch_view o_pivot oi oi-view-pivot"]').click()
    time.sleep(5)
    driver.find_element(By.XPATH, '//button[@class="btn btn-secondary fa fa-download o_pivot_download"]').click()
    time.sleep(20)


def transform_file():
    """Сохрание файла в формате csv."""   
    files = os.listdir(PATH_TO_XLSX)
    for file in files:
        if file.endswith('.xlsx'):
            book = load_workbook(os.path.join(PATH_TO_XLSX, file))
            sheet = book.active
            range_row = sheet.max_row
            continue

    data_quality = {'All rows': range_row,'dirty rows': 0,'clear_rows': 0} # Проверка качества данных.

    # Запись и чистка данных.
    with open(PATH_TO_CSV, 'a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file, delimiter=',')
        for i in range(1, range_row + 1):
            raw_data_col1 = f'{sheet['A' + str(i)].value}'
            raw_data_col2 = f'{sheet['B' + str(i)].value}'
            clear_col1 = raw_data_col1.strip()
            clear_col2 = raw_data_col2.strip()           
            if '[' not in clear_col1:
                data_quality['dirty rows'] += 1
                continue
            elif ',' in clear_col1:
               data_quality['clear_rows'] += 1
               del_comma = clear_col1.replace(',', '')
               writer.writerow([del_comma, clear_col2])
            else:
                data_quality['clear_rows'] += 1
                writer.writerow([clear_col1, clear_col2])
    
    # Запись логов в файл log.txt.
    current_dir = os.path.join(os.curdir, 'log.txt')
    with open (current_dir, 'a', newline='', encoding='utf-8') as file:
        current_time = dt.now()
        file.write(f'[DATA INFO] [{current_time.strftime("%d/%m/%Y, %H:%M:%S")}]\n')
        for key, values in data_quality.items():
            file.write(f'{key}: {values}\n')


if __name__=="__main__":
    load_dotenv()
    del_files()
    load_dataset()
    transform_file()

    
    