from common_variables import OWNER, LAUNCH_TIME, moscow_tz
from airflow.operators.python import PythonOperator
from airflow import DAG
from openpyxl import load_workbook
from datetime import datetime as dt
import os
import csv



"""DAG выполняет следующие действия:
   Трансформация файла .xls -> csv из источника.
   Запись логов в tmp_data/log.txt.
"""

# Импорт path.
PATH_TO_XLSX = '/opt/airflow/tmp_data'   
PATH_TO_CSV = '/opt/airflow/tmp_data/raw_data.csv' 
PATH_TO_LOG = '/opt/airflow/tmp_data/log.txt' 

DAG_ID = 'transform_file'


def transform_file():
    """Сохрание файла в формате csv."""   
    files = os.listdir(PATH_TO_XLSX)

    for file in files:
        if file.endswith('.xlsx'):
            book = load_workbook(os.path.join(PATH_TO_XLSX, file))
            sheet = book.active
            range_row = sheet.max_row
            continue

    data_quality = {'All rows': range_row,'dirty rows': 0,'clear_rows': 0} # Проверка качества данных.

    # Запись и чистка данных.
    with open(PATH_TO_CSV, 'a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file, delimiter=',')
        for i in range(1, range_row + 1):
            raw_data_col1 = f"{sheet['A' + str(i)].value}"
            raw_data_col2 = f"{sheet['B' + str(i)].value}"
            clear_col1 = raw_data_col1.strip()
            clear_col2 = raw_data_col2.strip()           
            if '[' not in clear_col1:
                data_quality['dirty rows'] += 1
                continue
            elif ',' in clear_col1:
               data_quality['clear_rows'] += 1
               del_comma = clear_col1.replace(',', '')
               writer.writerow([del_comma, clear_col2])
            else:
                data_quality['clear_rows'] += 1
                writer.writerow([clear_col1, clear_col2])
    
    # Запись логов в файл log.txt.
    with open (PATH_TO_LOG, 'a', newline='', encoding='utf-8') as file:
        current_time = dt.now(moscow_tz)
        file.write(f'[DATA INFO] [{current_time.strftime("%d/%m/%Y, %H:%M:%S")}]\n')
        for key, values in data_quality.items():
            file.write(f'{key}: {values}\n')



args = {
    'owner': OWNER,
    'start_date': dt(2025, 6, 4, tzinfo=moscow_tz)
}

with DAG(
    dag_id=DAG_ID,
    default_args=args,
    schedule_interval=LAUNCH_TIME,
    catchup=False,
    is_paused_upon_creation=True, 
    tags=['transform_file']
    
) as dag:
    
    task1 = PythonOperator(
        task_id='task1',
        python_callable=transform_file
    )

    
